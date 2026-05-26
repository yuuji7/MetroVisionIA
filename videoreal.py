# -*- coding: utf-8 -*-
"""
SAM 3.1 - Mapeador e Contabilizador de Câmera em Tempo Real (Zero-Shot)
====================================================================
Este utilitário permite capturar o feed da câmera do PC em tempo real, 
extrair frames em um ciclo temporizado configurável, aplicar a segmentação
com o modelo SAM 3.1 baseado em um prompt de texto e exibir a contagem
de objetos de forma contínua, com liberação rígida de memória a cada ciclo.
"""

import os
import gc
import sys
import time
import threading
import tkinter as tk
from tkinter import ttk, messagebox
import cv2
import numpy as np
import torch
from PIL import Image, ImageTk

# ============================================================
# PATCH DE COMPATIBILIDADE E OTIMIZAÇÃO DE SDPA PARA RTX 3060 LAPTOP
# ============================================================
import functools
import torch.nn.functional as _F_torch

# Ativação dos backends otimizados de SDPA com fallback seguro
try:
    torch.backends.cuda.enable_flash_sdp(True)
    torch.backends.cuda.enable_mem_efficient_sdp(True)
    torch.backends.cuda.enable_math_sdp(True)
    os.environ["PYTORCH_CUDA_ALLOC_CONF"] = "max_split_size_mb:128"
except Exception:
    pass

_original_sdpa = _F_torch.scaled_dot_product_attention

@functools.wraps(_original_sdpa)
def _safe_sdpa(query, key, value, attn_mask=None, dropout_p=0.0, is_causal=False, **kwargs):
    """SDPA com fallback dinâmico e automático para backend MATH se falhar."""
    try:
        return _original_sdpa(query, key, value, attn_mask=attn_mask,
                               dropout_p=dropout_p, is_causal=is_causal, **kwargs)
    except RuntimeError as e:
        err_str = str(e)
        if "No available kernel" in err_str or "not supported" in err_str.lower() or "abort" in err_str.lower():
            try:
                # Fallback temporário para MATH
                torch.backends.cuda.enable_flash_sdp(False)
                torch.backends.cuda.enable_mem_efficient_sdp(False)
                torch.backends.cuda.enable_math_sdp(True)
                res = _original_sdpa(query, key, value, attn_mask=attn_mask,
                                     dropout_p=dropout_p, is_causal=is_causal, **kwargs)
                # Restaura os kernels rápidos para as próximas chamadas
                torch.backends.cuda.enable_flash_sdp(True)
                torch.backends.cuda.enable_mem_efficient_sdp(True)
                return res
            except Exception:
                pass
        raise

# Patch em ambas as referências possíveis
_F_torch.scaled_dot_product_attention = _safe_sdpa
if hasattr(torch, "scaled_dot_product_attention"):
    torch.scaled_dot_product_attention = _safe_sdpa


# Carregamento de variáveis de ambiente (.env)
try:
    from dotenv import load_dotenv
    load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"))
except ImportError:
    pass  # python-dotenv não instalado; espera-se que as variáveis estejam no ambiente

# Adicionar caminhos para o SAM 3
_sam3_path = os.environ.get("SAM3_PATH", "")
if _sam3_path:
    sys.path.append(_sam3_path)
sys.path.append(os.path.join(os.path.dirname(os.path.abspath(__file__)), "src"))

try:
    from sam3 import build_sam3_image_model
    from sam3.model.sam3_image_processor import Sam3Processor
    SAM3_AVAILABLE = True
except ImportError as e:
    SAM3_AVAILABLE = False
    SAM3_IMPORT_ERROR = str(e)

# Cores de alto contraste para exibição das máscaras
COLORS = np.array([
    [0.9, 0.1, 0.1],  # Vermelho
    [0.1, 0.8, 0.1],  # Verde
    [0.1, 0.5, 0.95], # Azul
    [0.95, 0.85, 0.1],# Amarelo
    [0.9, 0.1, 0.9],  # Magenta
    [0.1, 0.9, 0.9],  # Ciano
    [0.95, 0.5, 0.1], # Laranja
    [0.55, 0.1, 0.95],# Roxo
    [0.1, 0.95, 0.5]  # Lima
])

class SAM3CameraSegmenterApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("SAM 3.1 - Segmentação de Câmera em Tempo Real & Contabilizador")
        self.geometry("1200x880")
        self.configure(bg="#1e1e1f")
        
        # Variáveis do Modelo
        self.model = None
        self.processor = None
        
        # Variáveis da Câmera e Origem
        self.cap = None
        self.camera_running = False
        self.last_raw_frame = None
        self.show_live_feed = True
        self.video_path = ""
        self.live_overlay_mask = None
        self.live_overlay_bool = None
        
        # Variáveis do Ciclo Temporizado
        self.cycle_running = False
        self.time_left = 5.0
        self.interval = 5.0
        self.is_processing = False
        
        # Estilos Dark Mode Premium
        self.style = ttk.Style()
        self.style.theme_use("clam")
        self.style.configure(".", background="#1e1e1f", foreground="#ffffff", fieldbackground="#2d2d2d")
        self.style.configure("TButton", background="#3c3c3d", foreground="#ffffff", borderwidth=0, padding=6, font=("Helvetica", 9, "bold"))
        self.style.map("TButton", background=[("active", "#4d4d4f")])
        self.style.configure("TLabel", background="#1e1e1f", foreground="#ffffff")
        self.style.configure("TEntry", fieldbackground="#2d2d2d", foreground="#ffffff", insertcolor="#ffffff")
        self.style.configure("TCombobox", fieldbackground="#2d2d2d", background="#3c3c3d", foreground="#ffffff")
        self.style.map("TCombobox", fieldbackground=[("readonly", "#2d2d2d")], foreground=[("readonly", "#ffffff")])
        
        self.create_widgets()
        
        # Carregamento assíncrono do modelo SAM 3.1
        if SAM3_AVAILABLE:
            self.log("[STATUS] Carregando pesos do SAM 3.1 Image Model...")
            threading.Thread(target=self.load_sam3_model, daemon=True).start()
        else:
            self.log(f"[ERRO CRÍTICO] Falha ao importar SAM 3.1: {SAM3_IMPORT_ERROR}")
            messagebox.showerror("Erro de Importação", f"Erro: {SAM3_IMPORT_ERROR}")
            
        # Tratamento de fechar janela
        self.protocol("WM_DELETE_WINDOW", self.on_closing)

    def create_widgets(self):
        # Painel Superior de Controles
        ctrl_frame = tk.Frame(self, bg="#252526", pady=10, padx=15, highlightbackground="#3c3c3c", highlightthickness=1)
        ctrl_frame.pack(fill="x", side="top", padx=10, pady=10)
        
        # Título
        tk.Label(ctrl_frame, text="PROJETO CONTADOR - SEGMENTAÇÃO DE CÂMERA EM TEMPO REAL", bg="#252526", fg="#ffaa00", font=("Helvetica", 11, "bold")).grid(row=0, column=0, columnspan=6, sticky="w", pady=(0, 10))
        
        # Linha 1: Fonte de Entrada e Ciclo
        tk.Label(ctrl_frame, text="Origem de Captura:", bg="#252526", fg="#ffffff", font=("Helvetica", 9, "bold")).grid(row=1, column=0, sticky="w", pady=5)
        self.combo_source = ttk.Combobox(ctrl_frame, values=["🎥 Câmera (Webcam)", "📁 Vídeo Local"], width=18, state="readonly")
        self.combo_source.set("🎥 Câmera (Webcam)")
        self.combo_source.grid(row=1, column=1, padx=5, pady=5, sticky="w")
        self.combo_source.bind("<<ComboboxSelected>>", self.on_source_changed)
        
        self.btn_camera = tk.Button(ctrl_frame, text="🎥 Abrir Câmera", bg="#28a745", fg="#ffffff", font=("Helvetica", 9, "bold"), activebackground="#218838", relief="flat", padx=10, command=self.toggle_camera)
        self.btn_camera.grid(row=1, column=2, padx=5, pady=5)
        
        tk.Label(ctrl_frame, text="Intervalo de Ciclo:", bg="#252526", fg="#ffffff", font=("Helvetica", 9, "bold")).grid(row=1, column=3, sticky="w", pady=5, padx=(15, 0))
        self.combo_interval = ttk.Combobox(ctrl_frame, values=["1", "2", "3", "5", "10", "15", "30", "60"], width=8)
        self.combo_interval.set("5")
        self.combo_interval.grid(row=1, column=4, padx=5, pady=5, sticky="w")
        tk.Label(ctrl_frame, text="segundos", bg="#252526", fg="#a0a0a0", font=("Helvetica", 9, "italic")).grid(row=1, column=5, sticky="w", pady=5)
        
        # Linha 2: Prompt e Confiança
        tk.Label(ctrl_frame, text="Prompt (Alvo):", bg="#252526", fg="#ffffff", font=("Helvetica", 9, "bold")).grid(row=2, column=0, sticky="w", pady=5)
        self.entry_prompt = ttk.Entry(ctrl_frame, width=21)
        self.entry_prompt.insert(0, "person")
        self.entry_prompt.grid(row=2, column=1, padx=5, pady=5, sticky="w")
        
        self.btn_cycle = tk.Button(ctrl_frame, text="⚡ Iniciar Ciclo de Captura", bg="#007acc", fg="#ffffff", font=("Helvetica", 9, "bold"), activebackground="#005999", relief="flat", state="disabled", padx=10, command=self.toggle_cycle)
        self.btn_cycle.grid(row=2, column=2, padx=5, pady=5)
        
        tk.Label(ctrl_frame, text="Limiar Confiança:", bg="#252526", fg="#ffffff", font=("Helvetica", 9, "bold")).grid(row=2, column=3, sticky="w", pady=5, padx=(15, 0))
        self.val_confidence = tk.DoubleVar(value=0.45)
        self.slider_conf = ttk.Scale(ctrl_frame, from_=0.1, to=0.9, variable=self.val_confidence, orient="horizontal", command=self.update_conf_label)
        self.slider_conf.grid(row=2, column=4, sticky="ew", padx=5, pady=5)
        self.lbl_conf = tk.Label(ctrl_frame, text="0.45", bg="#252526", fg="#a0a0a0", font=("Helvetica", 9, "bold"))
        self.lbl_conf.grid(row=2, column=5, sticky="w", padx=5, pady=5)
        
        # Linha 3: Resolução Customizada (Otimização Extrema de VRAM/Desempenho)
        tk.Label(ctrl_frame, text="Resolução (Altura px):", bg="#252526", fg="#00aaff", font=("Helvetica", 9, "bold")).grid(row=3, column=0, sticky="w", pady=5)
        self.entry_res_height = ttk.Entry(ctrl_frame, width=22)
        self.entry_res_height.insert(0, "240")  # Padrão 240px para velocidade máxima de contagem
        self.entry_res_height.grid(row=3, column=1, padx=5, pady=5, sticky="w")
        
        tk.Label(ctrl_frame, text="(Valores menores como 180, 240 ou 320 aceleram absurdamente a contagem e evitam estouros de memória)", bg="#252526", fg="#8e8e93", font=("Helvetica", 8, "italic")).grid(row=3, column=2, columnspan=4, sticky="w", padx=5, pady=5)
        
        # Linha 4: Seletor de Índice de Câmera e Arquivo de Vídeo Local
        tk.Label(ctrl_frame, text="Configuração Fonte:", bg="#252526", fg="#ffffff", font=("Helvetica", 9, "bold")).grid(row=4, column=0, sticky="w", pady=5)
        self.combo_cam_index = ttk.Combobox(ctrl_frame, values=["0 (Webcam Padrão)", "1", "2"], width=18, state="readonly")
        self.combo_cam_index.set("0 (Webcam Padrão)")
        self.combo_cam_index.grid(row=4, column=1, padx=5, pady=5, sticky="w")
        
        tk.Label(ctrl_frame, text="Arquivo de Vídeo:", bg="#252526", fg="#ffffff", font=("Helvetica", 9, "bold")).grid(row=4, column=2, sticky="w", pady=5, padx=(5, 0))
        self.entry_video_path = ttk.Entry(ctrl_frame, width=25, state="disabled")
        self.entry_video_path.grid(row=4, column=3, padx=5, pady=5, sticky="w")
        
        self.btn_select_video = ttk.Button(ctrl_frame, text="Selecionar...", state="disabled", command=self.select_video_file)
        self.btn_select_video.grid(row=4, column=4, padx=5, pady=5, sticky="w")
        
        # Painel Central: Canvas de Exibição
        viewer_container = tk.Frame(self, bg="#1e1e1f")
        viewer_container.pack(fill="both", expand=True, padx=10, pady=5)
        
        viewer_frame = ttk.LabelFrame(viewer_container, text="Feed da Câmera em Tempo Real (Mapeamento & Máscaras)", padding=5)
        viewer_frame.pack(fill="both", expand=True)
        
        self.canvas_image = tk.Canvas(viewer_frame, bg="#151515", highlightthickness=0)
        self.canvas_image.pack(fill="both", expand=True)
        
        # Overlay do Contador e Indicador de Ciclo
        info_frame = tk.Frame(self, bg="#1e1e1f")
        info_frame.pack(fill="x", padx=15, pady=5)
        
        self.lbl_counter = tk.Label(info_frame, text="Objetos Encontrados: 0", bg="#1e1e1f", fg="#00ff00", font=("Helvetica", 14, "bold"), anchor="w")
        self.lbl_counter.pack(side="left", fill="x", expand=True)
        
        self.lbl_timer = tk.Label(info_frame, text="Ciclo: Parado", bg="#1e1e1f", fg="#ffaa00", font=("Helvetica", 12, "bold"), anchor="e")
        self.lbl_timer.pack(side="right", padx=10)
        
        # Botão para abrir o arquivo Excel diretamente da interface
        self.btn_open_excel = tk.Button(info_frame, text="📊 Abrir Planilha Excel", bg="#107c41", fg="#ffffff", font=("Helvetica", 9, "bold"), activebackground="#0b592e", relief="flat", padx=10, command=self.open_excel_file)
        self.btn_open_excel.pack(side="right", padx=5)
        
        # Barra de Progresso do Processamento
        self.progress_bar = ttk.Progressbar(self, orient="horizontal", mode="indeterminate")
        self.progress_bar.pack(fill="x", padx=15, pady=5)
        
        # Console de status inferior
        self.log_frame = ttk.LabelFrame(self, text="Display de Processamento e Status de Memória", padding=5)
        self.log_frame.pack(fill="x", side="bottom", padx=10, pady=(0, 10))
        
        self.text_log = tk.Text(self.log_frame, wrap="word", height=6, bg="#101010", fg="#00ff66", insertbackground="white", state="disabled", font=("Courier", 9))
        self.text_log.pack(fill="both", expand=True)

    def log(self, msg):
        timestamp = time.strftime("[%H:%M:%S]")
        def _log():
            self.text_log.config(state="normal")
            self.text_log.insert("end", f"{timestamp} {msg}\n")
            self.text_log.config(state="disabled")
            self.text_log.see("end")
        self.after(0, _log)

    def update_conf_label(self, val):
        self.lbl_conf.config(text=f"{float(val):.2f}")

    def load_sam3_model(self):
        try:
            device = "cuda" if torch.cuda.is_available() else "cpu"
            self.log(f"[MODELO] Inicializando SAM 3.1 no dispositivo '{device.upper()}'...")
            
            # Carregar pesos
            self.model = build_sam3_image_model(compile=False, load_from_HF=True, device=device)
            self.processor = Sam3Processor(self.model, device=device)
            
            self.log("[MODELO] SAM 3.1 carregado com sucesso na memória!")
            self.after(0, lambda: self.btn_cycle.config(state="normal"))
        except Exception as e:
            self.log(f"[ERRO CRÍTICO] Falha ao carregar o SAM 3.1: {e}")
            messagebox.showerror("Erro de Carregamento", f"Não foi possível carregar o modelo SAM 3.1:\n{e}")

    def on_source_changed(self, event=None):
        source = self.combo_source.get()
        if "Câmera" in source:
            self.combo_cam_index.config(state="readonly")
            self.entry_video_path.config(state="disabled")
            self.btn_select_video.config(state="disabled")
            self.btn_camera.config(text="🎥 Abrir Câmera")
        else:
            self.combo_cam_index.config(state="disabled")
            self.entry_video_path.config(state="normal")
            self.btn_select_video.config(state="normal")
            self.btn_camera.config(text="▶ Iniciar Vídeo")

    def select_video_file(self):
        from tkinter import filedialog
        file = filedialog.askopenfilename(
            title="Selecionar Vídeo Local",
            filetypes=[("Vídeos", "*.mp4 *.avi *.mov *.mkv *.webm *.PNG *.jpg *.jpeg *.png *.JPG *.JPEG")]
        )
        if file:
            self.entry_video_path.delete(0, "end")
            self.entry_video_path.insert(0, file)
            self.log(f"[VÍDEO] Arquivo de vídeo selecionado: {os.path.basename(file)}")

    def toggle_camera(self):
        source = self.combo_source.get()
        is_video = "Vídeo" in source
        
        if self.camera_running:
            # Parar reprodução / câmera
            self.camera_running = False
            self.show_live_feed = False
            if self.cap is not None:
                self.cap.release()
                self.cap = None
            
            if is_video:
                self.btn_camera.config(text="▶ Iniciar Vídeo", bg="#28a745", activebackground="#218838")
                self.log("[VÍDEO] Reprodução interrompida e arquivo fechado.")
            else:
                self.btn_camera.config(text="🎥 Abrir Câmera", bg="#28a745", activebackground="#218838")
                self.log("[CÂMERA] Câmera fechada e recursos liberados.")
            self.canvas_image.delete("all")
        else:
            if is_video:
                # Iniciar Vídeo Local
                video_path = self.entry_video_path.get().strip()
                if not video_path:
                    messagebox.showwarning("Aviso", "Por favor, selecione um arquivo de vídeo local primeiro.")
                    return
                self.log(f"[VÍDEO] Abrindo arquivo de vídeo {os.path.basename(video_path)}...")
                self.btn_camera.config(text="Carregando...", state="disabled")
                self.update_idletasks()
                threading.Thread(target=self._init_video_thread, args=(video_path,), daemon=True).start()
            else:
                # Abrir câmera
                cam_str = self.combo_cam_index.get()
                try:
                    cam_idx = int(cam_str.split()[0])
                except Exception:
                    cam_idx = 0
                    
                self.log(f"[CÂMERA] Conectando ao dispositivo de captura {cam_idx}...")
                self.btn_camera.config(text="Conectando...", state="disabled")
                self.update_idletasks()
                
                # Inicialização assíncrona para não travar a GUI
                threading.Thread(target=self._init_camera_thread, args=(cam_idx,), daemon=True).start()

    def _init_camera_thread(self, cam_idx):
        cap = cv2.VideoCapture(cam_idx)
        if not cap.isOpened():
            self.log(f"[ERRO] Não foi possível abrir a câmera no índice {cam_idx}")
            self.after(0, lambda: messagebox.showerror("Erro de Câmera", f"Não foi possível abrir a câmera {cam_idx}."))
            self.after(0, lambda: self.btn_camera.config(text="🎥 Abrir Câmera", state="normal"))
            return
            
        # Definir parâmetros de captura padrão para 480p de entrada para maior rapidez
        cap.set(cv2.CAP_PROP_FRAME_WIDTH, 854)
        cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
        
        self.cap = cap
        self.camera_running = True
        self.show_live_feed = True
        self.after(0, lambda: self.btn_camera.config(text="❌ Fechar Câmera", bg="#dc3545", activebackground="#bd2130", state="normal"))
        self.log(f"[CÂMERA] Câmera {cam_idx} aberta com sucesso! Feed em tempo real ativo.")
        
        # Iniciar loop de atualização de quadros
        self.after(30, self.update_camera_loop)

    def _init_video_thread(self, video_path):
        if not os.path.exists(video_path):
            self.log(f"[ERRO] Arquivo de vídeo não encontrado: {video_path}")
            self.after(0, lambda: messagebox.showerror("Erro de Arquivo", "Arquivo de vídeo não encontrado."))
            self.after(0, lambda: self.btn_camera.config(text="▶ Iniciar Vídeo", state="normal"))
            return
            
        cap = cv2.VideoCapture(video_path)
        if not cap.isOpened():
            self.log(f"[ERRO] Não foi possível abrir o arquivo de vídeo: {video_path}")
            self.after(0, lambda: messagebox.showerror("Erro de Vídeo", "Não foi possível abrir o arquivo de vídeo."))
            self.after(0, lambda: self.btn_camera.config(text="▶ Iniciar Vídeo", state="normal"))
            return
            
        self.cap = cap
        self.camera_running = True
        self.show_live_feed = True
        self.after(0, lambda: self.btn_camera.config(text="❌ Parar Vídeo", bg="#dc3545", activebackground="#bd2130", state="normal"))
        self.log(f"[VÍDEO] Arquivo de vídeo {os.path.basename(video_path)} aberto com sucesso! Feed de vídeo ativo.")
        
        # Iniciar loop de atualização de quadros
        self.after(30, self.update_camera_loop)

    def update_camera_loop(self):
        if not self.camera_running or self.cap is None:
            return
            
        ret, frame = self.cap.read()
        if ret:
            self.last_raw_frame = frame.copy()
            if self.show_live_feed:
                self.display_opencv_frame(frame)
        else:
            # Se for vídeo local, reiniciar do início (loop contínuo)
            source = self.combo_source.get()
            if "Vídeo" in source and self.cap is not None:
                self.cap.set(cv2.CAP_PROP_POS_FRAMES, 0)
                
        # Mantém atualizando a cada 30ms (~33 FPS)
        self.after(30, self.update_camera_loop)

    def display_opencv_frame(self, frame):
        # Converte de BGR (OpenCV) para RGB (PIL)
        rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        pil_img = Image.fromarray(rgb_frame)
        self.display_image(pil_img)

    def display_image(self, pil_img):
        self.canvas_image.delete("all")
        c_w = self.canvas_image.winfo_width()
        c_h = self.canvas_image.winfo_height()
        if c_w <= 1: c_w = 640
        if c_h <= 1: c_h = 480
        
        w, h = pil_img.size
        ratio = min(c_w / w, c_h / h)
        new_w = int(w * ratio)
        new_h = int(h * ratio)
        
        resized = pil_img.resize((new_w, new_h), Image.Resampling.LANCZOS)
        img_tk = ImageTk.PhotoImage(resized)
        
        # Manter referência para evitar Garbage Collection do Tkinter
        self._tk_img_show = img_tk
        self.canvas_image.create_image(c_w // 2, c_h // 2, anchor="center", image=img_tk)

    def toggle_cycle(self):
        if self.cycle_running:
            # Parar Ciclo
            self.cycle_running = False
            self.btn_cycle.config(text="⚡ Iniciar Ciclo de Captura", bg="#007acc", activebackground="#005999")
            self.lbl_timer.config(text="Ciclo: Parado", fg="#ffaa00")
            self.log("[CICLO] Ciclo temporizado interrompido pelo usuário.")
        else:
            # Validar e Iniciar Ciclo
            if not self.camera_running:
                messagebox.showwarning("Aviso", "Por favor, abra a câmera antes de iniciar o ciclo de capturas.")
                return
            if self.last_raw_frame is None:
                messagebox.showwarning("Aviso", "Aguarde até obter o primeiro quadro da câmera.")
                return
                
            try:
                self.interval = float(self.combo_interval.get().strip())
                if self.interval < 1.0:
                    raise ValueError("O intervalo deve ser de pelo menos 1 segundo.")
            except ValueError as e:
                messagebox.showerror("Erro de Parâmetro", f"Intervalo inválido: {e}")
                return
                
            self.cycle_running = True
            self.time_left = self.interval
            self.btn_cycle.config(text="🛑 Parar Captura de Ciclo", bg="#ff9900", activebackground="#cc7a00")
            self.log(f"[CICLO] Iniciando ciclo automático a cada {self.interval:.1f}s.")
            
            # Iniciar relógio de contagem regressiva
            self.after(100, self.tick_countdown)

    def tick_countdown(self):
        if not self.cycle_running:
            return
            
        if self.is_processing:
            # Se está processando o SAM 3.1, pausa o cronômetro visual
            self.lbl_timer.config(text="Processando...", fg="#00aaff")
            self.after(100, self.tick_countdown)
            return
            
        self.time_left -= 0.1
        if self.time_left <= 0:
            self.time_left = self.interval
            # Disparar Captura e Processamento
            self.is_processing = True
            self.progress_bar.start(10)
            self.lbl_counter.config(text="Processando frame...")
            
            # Capturar frame em tempo real SEM congelar ou pausar a exibição
            snapshot_frame = self.last_raw_frame.copy()
            prompt = self.entry_prompt.get().strip()
            
            # Obter a posição de tempo do vídeo no momento EXATO do print!
            video_time = self.get_video_timestamp()
            
            self.log(f"[CICLO] Ciclo atingido! Capturando frame em {video_time}...")
            threading.Thread(target=self._process_snapshot_thread, args=(snapshot_frame, prompt, video_time), daemon=True).start()
        else:
            self.lbl_timer.config(text=f"Próxima Captura em: {self.time_left:.1f}s", fg="#00ff66")
            
        self.after(100, self.tick_countdown)

    def _process_snapshot_thread(self, frame, prompt, video_time):
        try:
            device = "cuda" if torch.cuda.is_available() else "cpu"
            start_time = time.time()
            
            # 1. Converter frame de OpenCV BGR para PIL RGB
            cv_img = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            orig_img = Image.fromarray(cv_img)
            w, h = orig_img.size
            working_img = orig_img
            
            # 2. Redimensionamento Inteligente Customizado (Desempenho Máximo)
            try:
                target_h = int(self.entry_res_height.get().strip())
                if target_h < 32:
                    target_h = 32
                elif target_h > 2160:
                    target_h = 2160
            except ValueError:
                target_h = 240  # Fallback padrão ultra rápido
                
            target_w = int(w * (target_h / h))
            target_w = (target_w // 2) * 2
            target_h = (target_h // 2) * 2
            
            self.log(f"[RESOLUÇÃO] Ajustando imagem de processamento para {target_w}x{target_h} para otimizar velocidade...")
            working_img = orig_img.resize((target_w, target_h), Image.Resampling.LANCZOS)
                
            user_thresh = self.val_confidence.get()
            
            # 3. Executar Inferência SAM 3.1
            state = {}
            self.processor.set_confidence_threshold(user_thresh, state)
            
            with torch.inference_mode():
                if device == "cuda":
                    with torch.autocast("cuda", dtype=torch.bfloat16):
                        state = self.processor.set_image(working_img, state)
                        state = self.processor.set_text_prompt(prompt, state)
                else:
                    state = self.processor.set_image(working_img, state)
                    state = self.processor.set_text_prompt(prompt, state)
                    
            elapsed = time.time() - start_time
            
            # 4. Extrair Resultados
            raw_scores = state.get("scores")
            if raw_scores is not None:
                scores_np = raw_scores.cpu().float().numpy() if hasattr(raw_scores, "cpu") else np.array(raw_scores)
            else:
                scores_np = np.array([])
                
            num_detected = len(scores_np)
            keep_indices = np.arange(num_detected)
            
            self.log(f"[MAPEAMENTO] Concluído em {elapsed:.2f}s! Mapeados: {num_detected} '{prompt}'(s) (Confiança > {user_thresh:.2f}).")
            self.after(0, lambda: self.lbl_counter.config(text=f"Objetos Encontrados ({prompt}): {num_detected}"))
            
            # 5. Desenhar Máscaras sobre o print capturado (Sem interferir no vídeo) e preparar print mascarado
            masked_pil = orig_img
            if num_detected > 0:
                raw_masks = state.get("masks")
                if isinstance(raw_masks, torch.Tensor):
                    filtered_masks = raw_masks[keep_indices]
                else:
                    filtered_masks = [raw_masks[i] for i in keep_indices]
                
                # Gerar a imagem estática com máscaras e caixas para salvar em disco
                raw_boxes = state.get("boxes")
                raw_scores = state.get("scores")
                if isinstance(raw_masks, torch.Tensor):
                    filtered_boxes = raw_boxes[keep_indices]
                    filtered_scores = raw_scores[keep_indices]
                else:
                    filtered_boxes = [raw_boxes[i] for i in keep_indices]
                    filtered_scores = [raw_scores[i] for i in keep_indices]
                masked_pil = self.draw_results(orig_img, filtered_masks, filtered_boxes, filtered_scores)
                
            # Salvar print com máscara em disco para auditoria
            captures_dir = os.path.join("media", "capturas")
            os.makedirs(captures_dir, exist_ok=True)
            timestamp_file = time.strftime("%Y%m%d_%H%M%S")
            video_time_clean = video_time.replace(":", "").replace(".", "").replace(" ", "_")
            img_filename = f"cap_{timestamp_file}_vt{video_time_clean}_{prompt}.png"
            img_path = os.path.join(captures_dir, img_filename)
            masked_pil.save(img_path)
            self.log(f"[CAPTURA] Print mascarado salvo em '{img_path}'")
            
            # 6. Salvar Informações no Excel (Horário, Tempo do Vídeo, Quantidade, Objeto, Caminho + Imagem Embutida)
            self.save_to_excel(num_detected, prompt, video_time, img_path)
            
            # ============================================================
            # LIMPEZA ESTRITA DE MEMÓRIA (Evita travamentos e atraso de ciclo)
            # ============================================================
            state.clear()
            del state
            
            # Forçar coleta imediata de lixo do Python
            gc.collect()
            
            # Se rodando na GPU CUDA, limpa o cache de VRAM
            if device == "cuda":
                torch.cuda.empty_cache()
                
            self.log("[MEMÓRIA] Coleta de lixo executada. RAM/VRAM liberada com sucesso.")
            
        except Exception as e:
            self.log(f"[ERRO CRÍTICO] Falha no pipeline de segmentação: {e}")
        finally:
            # Retomar fluxo normal e interromper barra
            self.is_processing = False
            self.after(0, self.progress_bar.stop)

    def draw_results(self, original_img, masks, boxes, scores):
        # Converte a imagem PIL para um array numpy (BGR) para processamento rápido
        frame = cv2.cvtColor(np.array(original_img), cv2.COLOR_RGB2BGR)
        h, w, _ = frame.shape
        
        overlay_mask = np.zeros_like(frame, dtype=np.uint8)
        overlay_mask_bool = np.zeros((h, w), dtype=bool)
        
        if hasattr(masks, "cpu"):
            masks = masks.cpu().numpy()
        if hasattr(boxes, "cpu"):
            boxes = boxes.cpu().float().numpy()
            
        num_detected = len(masks)
        for idx in range(num_detected):
            mask = masks[idx]
            if mask.ndim == 3:
                mask = mask[0]
                
            if mask.shape != (h, w):
                mask = cv2.resize(mask.astype(np.uint8), (w, h), interpolation=cv2.INTER_NEAREST).astype(bool)
                
            rgb = COLORS[idx % len(COLORS)]
            color_bgr = (int(rgb[2] * 255), int(rgb[1] * 255), int(rgb[0] * 255))
            
            overlay_mask[mask] = color_bgr
            overlay_mask_bool[mask] = True
            
        # Aplica efeito Alpha blending (60% original, 40% cor da máscara)
        if num_detected > 0:
            frame[overlay_mask_bool] = (frame[overlay_mask_bool] * 0.6 + overlay_mask[overlay_mask_bool] * 0.4).astype(np.uint8)
            
        return Image.fromarray(cv2.cvtColor(frame, cv2.COLOR_BGR2RGB))


    def save_to_excel(self, count, prompt, video_time, img_path):
        excel_path = "relatorio_contagem.xlsx"
        timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
        
        import queue
        saved = False
        
        while not saved:
            try:
                import openpyxl
                from openpyxl import Workbook
                from openpyxl.drawing.image import Image as OpenpyxlImage
                
                if os.path.exists(excel_path):
                    wb = openpyxl.load_workbook(excel_path)
                    ws = wb.active
                else:
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Relatório de Contagens"
                    # Escrever cabeçalho se o arquivo for novo
                    ws.append(["Horário", "Tempo do Vídeo", "Quantidade Identificada", "Objeto Identificado", "Caminho da Imagem"])
                    
                    # Configurar larguras de coluna
                    ws.column_dimensions['A'].width = 20
                    ws.column_dimensions['B'].width = 25
                    ws.column_dimensions['C'].width = 25
                    ws.column_dimensions['D'].width = 25
                    ws.column_dimensions['E'].width = 45
                    ws.column_dimensions['F'].width = 25
                    
                # Adicionar nova linha
                row_data = [timestamp, video_time, count, prompt, img_path]
                ws.append(row_data)
                
                # Ajustar altura da linha e embutir imagem
                row_idx = ws.max_row
                ws.row_dimensions[row_idx].height = 80
                
                if os.path.exists(img_path):
                    img = OpenpyxlImage(img_path)
                    # Redimensionar para thumbnail de 150x100px
                    img.width = 150
                    img.height = 100
                    ws.add_image(img, f"F{row_idx}")
                    
                # Tenta salvar o arquivo Excel
                wb.save(excel_path)
                wb.close()
                saved = True
                self.log(f"[EXCEL] Dados salvos em '{excel_path}': {timestamp} | Tempo do Vídeo: {video_time} | Qtd: {count} | Imagem embutida na coluna F!")
                
            except PermissionError:
                # O arquivo está aberto no Excel e bloqueado pelo Windows!
                self.log(f"[AVISO EXCEL] O arquivo '{excel_path}' está bloqueado (aberto no Excel). Solicitando fechamento...")
                
                # Criar fila thread-safe para obter a resposta do diálogo executado na thread de GUI
                q = queue.Queue()
                
                def ask_user():
                    res = messagebox.askretrycancel(
                        "Planilha Bloqueada / Aberta",
                        f"A planilha '{excel_path}' está aberta no Microsoft Excel.\n\n"
                        "Por favor, FECHE a planilha no Excel e clique em 'Tentar Novamente' para salvar os dados capturados do ciclo.\n\n"
                        "Se você clicar em 'Cancelar', os dados deste ciclo serão descartados na planilha para evitar o travamento do aplicativo."
                    )
                    q.put(res)
                    
                # Agendar execução do diálogo na thread principal
                self.after(0, ask_user)
                
                # Bloquear thread em background esperando pela resposta do usuário
                retry = q.get()
                
                if not retry:
                    self.log(f"[ERRO EXCEL] Gravação cancelada pelo usuário. Os dados deste ciclo foram descartados da planilha.")
                    break
                    
            except Exception as e:
                self.log(f"[ERRO EXCEL] Falha crítica ao gravar na planilha: {e}")
                break

    def get_video_timestamp(self):
        source = self.combo_source.get()
        if "Vídeo" in source and self.cap is not None:
            msec = self.cap.get(cv2.CAP_PROP_POS_MSEC)
            total_seconds = int(msec / 1000)
            milliseconds = int(msec % 1000)
            hours = total_seconds // 3600
            minutes = (total_seconds % 3600) // 60
            seconds = total_seconds % 60
            return f"{hours:02d}:{minutes:02d}:{seconds:02d}.{milliseconds:03d}"
        else:
            return "Tempo Real (Câmera)"

    def open_excel_file(self):
        excel_path = "relatorio_contagem.xlsx"
        if os.path.exists(excel_path):
            try:
                os.startfile(excel_path)
                self.log(f"[EXCEL] Abrindo arquivo de planilha '{excel_path}'...")
            except Exception as e:
                messagebox.showerror("Erro ao Abrir", f"Não foi possível abrir o Excel:\n{e}")
        else:
            messagebox.showinfo("Aviso", "O arquivo de relatório do Excel ainda não foi criado. Inicie uma captura de ciclo para registrar dados.")

    def on_closing(self):
        self.camera_running = False
        self.cycle_running = False
        if self.cap is not None:
            self.cap.release()
        self.destroy()

if __name__ == "__main__":
    app = SAM3CameraSegmenterApp()
    app.mainloop()

# -*- coding: utf-8 -*-
"""
Utilitários de integração e gravação de relatórios em planilhas Excel.
"""
import os
import time

def export_to_excel(count, prompt, video_time, img_path, excel_path="relatorio_contagem.xlsx", log_fn=print, prompt_retry_fn=None):
    """
    Grava os dados de contagem em uma planilha Excel e insere uma thumbnail correspondente.
    Trata casos de bloqueio do arquivo por abertura simultânea no Excel.
    
    Parâmetros:
    - count: Número de objetos identificados.
    - prompt: Alvo da busca (e.g. "person").
    - video_time: Timestamp relativo no vídeo ou "Tempo Real (Câmera)".
    - img_path: Caminho no disco para a imagem capturada e mascarada.
    - excel_path: Caminho do arquivo Excel de destino.
    - log_fn: Função callback para logging de status.
    - prompt_retry_fn: Função callback chamada em caso de erro de permissão (bloqueio do Excel)
                       para perguntar ao usuário se ele quer tentar novamente.
                       Assinatura esperada: prompt_retry_fn(excel_path) -> bool (True para re-tentar).
    """
    timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
    saved = False
    
    while not saved:
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.drawing.image import Image as OpenpyxlImage
            
            if os.path.exists(excel_path):
                wb = openpyxl.load_workbook(excel_path)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Relatório de Contagens"
                # Escrever cabeçalho se o arquivo for novo
                ws.append(["Horário", "Tempo do Vídeo", "Quantidade Identificada", "Objeto Identificado", "Caminho da Imagem"])
                
                # Configurar larguras de coluna
                ws.column_dimensions['A'].width = 20
                ws.column_dimensions['B'].width = 25
                ws.column_dimensions['C'].width = 25
                ws.column_dimensions['D'].width = 25
                ws.column_dimensions['E'].width = 45
                ws.column_dimensions['F'].width = 25
                
            # Adicionar nova linha
            row_data = [timestamp, video_time, count, prompt, img_path]
            ws.append(row_data)
            
            # Ajustar altura da linha e embutir imagem
            row_idx = ws.max_row
            ws.row_dimensions[row_idx].height = 80
            
            if os.path.exists(img_path):
                img = OpenpyxlImage(img_path)
                # Redimensionar para thumbnail de 150x100px
                img.width = 150
                img.height = 100
                ws.add_image(img, f"F{row_idx}")
                
            # Tenta salvar o arquivo Excel
            wb.save(excel_path)
            wb.close()
            saved = True
            log_fn(f"[EXCEL] Dados salvos em '{excel_path}': {timestamp} | Tempo do Vídeo: {video_time} | Qtd: {count} | Imagem embutida na coluna F!")
            return True
            
        except PermissionError:
            # O arquivo está aberto no Excel e bloqueado pelo Windows!
            log_fn(f"[AVISO EXCEL] O arquivo '{excel_path}' está bloqueado (aberto no Excel). Solicitando fechamento...")
            
            if prompt_retry_fn:
                retry = prompt_retry_fn(excel_path)
                if not retry:
                    log_fn(f"[ERRO EXCEL] Gravação cancelada pelo usuário. Os dados deste ciclo foram descartados da planilha.")
                    break
            else:
                log_fn(f"[ERRO EXCEL] Permissão negada para gravar em '{excel_path}' (arquivo possivelmente aberto). Gravação cancelada.")
                break
                
        except Exception as e:
            log_fn(f"[ERRO EXCEL] Falha crítica ao gravar na planilha: {e}")
            break
            
    return False
